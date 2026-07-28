import os
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def read_excel_file(file_path: str, read_formulas: bool = True) -> str:
    """
    Reads an Excel (.xlsx) file and returns its sheets formatted as markdown tables.
    If read_formulas is True, reads formulas (e.g. '=SUM(A1:A10)') instead of cached values.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    # load workbook
    wb = openpyxl.load_workbook(file_path, data_only=not read_formulas)
    
    markdown_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        markdown_parts.append(f"### Sheet: {sheet_name}\n")
        
        # Read all rows
        rows = list(ws.iter_rows(values_only=True))
        if not rows or all(all(cell is None for cell in row) for row in rows):
            markdown_parts.append("*This sheet is empty.*\n")
            continue
            
        # Determine maximum columns
        max_cols = max(len(row) for row in rows)
        
        # Convert row values to strings and format as markdown
        table_rows = []
        for r in rows:
            row_vals = []
            for cell in r:
                if cell is None:
                    row_vals.append("")
                else:
                    # Clean up newlines for markdown table compatibility
                    row_vals.append(str(cell).strip().replace('\n', ' '))
            # Fill missing trailing columns
            row_vals += [""] * (max_cols - len(row_vals))
            table_rows.append(row_vals)
            
        # Create markdown table
        # Header
        header = table_rows[0]
        markdown_parts.append("| " + " | ".join(header) + " |")
        markdown_parts.append("| " + " | ".join(["---"] * max_cols) + " |")
        
        # Data
        for row in table_rows[1:]:
            markdown_parts.append("| " + " | ".join(row) + " |")
            
        markdown_parts.append("") # blank line between sheets
        
    return "\n".join(markdown_parts)


def write_excel_file(file_path: str, sheets_data: Dict[str, List[List[Any]]]) -> str:
    """
    Creates or overwrites an Excel (.xlsx) file.
    sheets_data: Dictionary mapping sheet name to a 2D list of row values:
    {
        "Overview": [
            ["Metric", "Value"],
            ["Total Revenue", 15000],
            ["Expenses", 9000],
            ["Net Profit", "=B2-B3"]  # Formulas are supported!
        ]
    }
    """
    wb = openpyxl.Workbook()
    # Remove the default sheet that openpyxl creates automatically
    wb.remove(wb.active)
    
    for sheet_name, rows in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        
        if not rows:
            continue
            
        # Write rows
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
                
        # Premium Styling setup
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Classic Navy/Slate Blue
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        normal_font = Font(name="Segoe UI", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Set grid line visibility
        ws.views.sheetView[0].showGridLines = True
        
        # Determine columns widths dynamically and apply cell styles
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                cell.font = normal_font
                cell.border = thin_border
                
                # Check value length
                val_str = str(cell.value or '')
                if val_str.startswith('='):
                    # Default sizing for formulas
                    max_len = max(max_len, 10)
                else:
                    max_len = max(max_len, len(val_str))
                    
            # Auto-fit width with safety margin
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        # Apply header styling on first row
        for c_idx in range(1, len(rows[0]) + 1):
            cell = ws.cell(row=1, column=c_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
    # Create directory if it doesn't exist
    dir_name = os.path.dirname(file_path)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name)
        
    wb.save(file_path)
    return f"Excel workbook successfully written to {file_path}"
